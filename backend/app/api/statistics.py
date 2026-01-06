import smtplib
from typing import List

from io import BytesIO

from datetime import date, datetime

from email import encoders
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication

from flask import request, abort, make_response, jsonify
from sqlalchemy.sql.functions import concat
from sqlalchemy.types import Date

#from flask_mail import Message

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from .. import myfilters, mail
from ..models import MEAL_LUNCH, MEAL_SUPPER
from ..models import Staff, CheckInHistory, Reservation, Recipe, Combo
from ..utils import get_week_days, get_days
from ..status import STATUS_EMAIL_NOT_FOUND, STATUS_REQUIRED_ARG_INVALID, MESSAGES
from ..myrequest import get_staff

from . import api
from .base import admin_required

@api.route('/statistics/togo', methods=['POST'])
@admin_required
def togo_records():
    today = datetime.now()
    if today.hour > 15: # 下午3点之后
        mode = MEAL_SUPPER
    else:
        mode = MEAL_LUNCH
    togos = CheckInHistory.query.filter(CheckInHistory.occur_time.cast(Date) == today.date(),
                                        CheckInHistory.mode.op('&')(mode) == mode,
                                        CheckInHistory.togo_status.op('&')(0xFF) > 0)\
                                .order_by(CheckInHistory.togo_status, CheckInHistory.id).all()
    result = []
    for togo in togos:
        tg = togo.to_json()
        tg['staff'] = togo.staff.to_json()
        result.append(tg)

    return jsonify(result), 200

@api.route('/statistics/checkin/preview', methods=['POST'])
@admin_required
def checkin_records_preview():
    data = request.json

    row_header, records, total_checkin, total_person = checkin_records(*parse_data_from_client(data))
    return jsonify({'total_checkin': total_checkin, 'total_person': total_person, 'records': records, 'header': row_header}), 200

@api.route('/statistics/checkin/email', methods=['POST'])
@admin_required
def checkin_records_email():
    data = request.json
    staff = get_staff(request.headers.get('X-ACCESS-TOKEN'))
    if not staff.email:
        abort(make_response(jsonify(errcode=STATUS_EMAIL_NOT_FOUND, message=MESSAGES[STATUS_EMAIL_NOT_FOUND].format(staff.nickname)), 404))
    #row_header, records, total_checkin, total_person = parse_data_from_client(data)
    row_header, records, total_checkin, total_person = checkin_records(*parse_data_from_client(data))
    ## to generate excel file
    bytesIO = BytesIO()
    excel_checkin_record(bytesIO, row_header, records, [['签到总人数', total_person, '总人数', len(records)], ['签到总次数', total_checkin]])
    """
    message = Message(subject='签到统计表', sender='warmlab@outlook.com', recipients=[staff.email])
    if len(row_header) > 3:
        message.body = "您好，附件中包含了从{}到{}日的签到信息。".format(row_header[3], row_header[-1])
    else:
        message.body = "您好，附件中包含了一些签到信息。"
    message.attach("checkin.xlsx", "application/vnd.ms-excel", bytesIO.getvalue())
    mail.send(message)
    """

    message = MIMEMultipart()
    message['From'] = 'axu307@gmail.com'
    message['To'] =  staff.email
    message['Subject'] = '签到统计表'
    if len(row_header) > 3:
        message.attach(MIMEText("您好，附件中包含了从{}到{}日的签到信息。".format(row_header[3], row_header[-1]), 'plain'))
    else:
        message.attach(MIMEText("您好，附件中包含了一些签到信息。".format(row_header[3], row_header[-1]), 'plain'))

    part = MIMEBase('application', "octet-stream")
    part.set_payload(bytesIO.getvalue(), charset='utf-8')
    #part.set_payload(bytesIO.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition',
                    'attachment; filename="checkin.xlsx"')
    #attachement = MIMEApplication(bytesIO.getvalue(), name='checkin.xlsx')
    #attachement['Content-Disposition'] = 'attachment; filename="checkin.xlsx"'

    #message.attach(attachement)
    message.attach(part)
    if send_email(message, staff.email):
        return jsonify({}), 200
    else:
        return jsonify({}), 422 # Unprocessable Content

def parse_data_from_client(data):
    #row_header = ['seq', 'name']
    #row_header = ['姓名', '公司']
    name = data.get('name')
    company_ids = data.get('companies')
    time_period = []
    time_period_option = data.get('time_period_option')
    try:
        time_period_option = int(time_period_option)
    except Exception as e:
        time_period_option = 1

    if time_period_option == 1 or time_period_option == 2:
        time_period = get_week_days(1-time_period_option)
        #row_header.extend(myfilters.week_short_name_list)
    else:
        begin_date = data.get('begin_date')
        end_date = data.get('end_date')
        if not begin_date or not end_date:
            time_period = []
        else:
            print(begin_date, end_date)
            try:
                begin_date = datetime.strptime(begin_date, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            except Exception as e:
                abort(make_response(jsonify(errcode=STATUS_REQUIRED_ARG_INVALID, message=MESSAGES[STATUS_REQUIRED_ARG_INVALID]), 400))

            time_period = get_days(begin_date, end_date)

    return name, company_ids, time_period
            #row_header.extend(time_period)
    
    #company_id = data.get('company_id')
    #try:
    #    company_id = int(company_id)
    #except Exception as e:
    #    company_id = 0

def excel_checkin_record(bytesIO, row_header, records, summaries):
    wb = Workbook()
    # grab the active worksheet
    ws = wb.active
    ws['A1'] = "从{}日到{}日的签到信息".format(row_header[3], row_header[-1])
    ws.append(row_header)
    for index, record in enumerate(records, start=1):
        a = [index, record['last_name'] + record['first_name'], record['company']]
        a.extend(record['days'])
        ws.append(a)

    last_column = ord('A')+len(row_header)-1
    last_row = len(records) + 2

    tab = Table(displayName="CheckinTable", ref="A2:{}{}".format(chr(last_column), last_row))
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    ws.append([])
    #last_line = ord(last_line) + 1
    for summary in summaries:
        ws.append(summary)

    wb.save(bytesIO)

def checkin_records(name: str, company_ids: List[int], time_period: List[date]):
    total_person = 0
    total_checkin = 0
    row_header = ['#', '姓名', '公司']
    #row_header.extend(time_period)
    for day in time_period:
        row_header.append(day.strftime('%m/%d'))

    if not company_ids or not time_period:
        return [], [], 0, 0

    if not name or not name.strip():
        staffs = Staff.query.filter(Staff.company_id.in_(company_ids), Staff.is_active==True)
    else:
        staffs = Staff.query.filter(concat(Staff.last_name, Staff.first_name).like('%{}%'.format(name.strip())),
                                    Staff.company_id.in_(company_ids), Staff.is_active==True)

    records = []
    for staff in staffs:
        histories = CheckInHistory.query\
                        .filter(CheckInHistory.staff_id==staff.id, CheckInHistory.occur_time >= time_period[0],
                                CheckInHistory.occur_time <= time_period[-1]).all()
        date_modes = {history.occur_time.date():history.mode for history in histories}
        dates = {}
        for history in histories:
            d = history.occur_time.date()
            if d not in dates:
                dates[d] = history.mode
                total_checkin += 1
            else:
                dates[d] |= history.mode

        record = {'last_name': staff.last_name, 'first_name': staff.first_name, 'company': staff.company.name}
        days = []
        any_checkedin = False
        for day in time_period:
            if day in dates:
                days.append(date_modes[day])
                any_checkedin = True
            else:
                days.append(0)
        if any_checkedin:
            total_person += 1
        record['days'] = days

        records.append(record)

    return row_header, records, total_checkin, total_person

@api.route('/statistics/reservation/preview', methods=['POST'])
@admin_required
def reservation_records_preview():
    data = request.json
    rs = reservation_records(*parse_data_from_client(data))

    return jsonify(rs), 200

def reservation_records(name: str, company_ids: List[int], time_period: List[date]):
    total_person = 0
    total_reservation = 0
    total_pickup = 0
    if not name or not name.strip():
        staffs = Staff.query.filter(Staff.company_id.in_(company_ids), Staff.is_active==True)
    else:
        staffs = Staff.query.filter(concat(Staff.last_name, Staff.first_name).like('%{}%'.format(name.strip())),
                                    Staff.company_id.in_(company_ids), Staff.is_active==True)

    total_combo = {}

    records = []
    for staff in staffs:
        reservations = Reservation.query\
                        .filter(Reservation.staff_id==staff.id, Reservation.occur_time >= time_period[0],
                                Reservation.occur_time <= time_period[-1]).all()
        if reservations:
            total_reservation += len(reservations)
            total_person += 1

            rs = []
            for reservation in reservations:
                tmp = reservation.to_json()
                tmp['combo'] = reservation.combo.to_json()
                if not total_combo.get(reservation.combo.id):
                    total_combo[reservation.combo.id] = 1
                else:
                    total_combo[reservation.combo.id] += 1
                rs.append(tmp)
                if reservation.pickup_time:
                    total_pickup += 1
        
            r = {'staff': staff.to_json(), 'reservations': rs}
        else:
            r = {'staff': staff.to_json(), 'reservations': []}
        records.append(r)

        cs = []
        combos = Combo.query.all()
        for c in combos:
            print(c)
            cj = c.to_json()
            if not total_combo.get(c.id):
                cj['total'] = 0
            else:
                cj['total'] = total_combo[c.id]
            cs.append(cj)

    return {'begin_date': time_period[0], 'end_date': time_period[-1],
            'total_person': total_person, 'total_reservation': total_reservation, 'total_pickup': total_pickup,
            'combos': cs, 'records': records}

@api.route('/statistics/reservation/email', methods=['POST'])
@admin_required
def reservation_records_email():
    data = request.json
    staff = get_staff(request.headers.get('X-ACCESS-TOKEN'))
    if not staff.email:
        abort(make_response(jsonify(errcode=STATUS_EMAIL_NOT_FOUND, message=MESSAGES[STATUS_EMAIL_NOT_FOUND].format(staff.nickname)), 404))
    #row_header, records, total_checkin, total_person = parse_data_from_client(data)
    records = reservation_records(*parse_data_from_client(data))
    ## to generate excel file
    row_header = ['#', '姓名', '公司', '套餐类型', '预约时间', '领取时间']
    bytesIO = BytesIO()
    #excel_reservation_record(bytesIO, row_header, records, [['预约总人数', total_person, '总人数', len(records)], ['预约总次数', total_checkin]])
    excel_reservation_record(bytesIO, row_header, records)
    bytesIO.seek(0)
    # message = Message(subject='预约领取统计表', sender='warmlab@outlook.com', recipients=[staff.email])
    message = MIMEMultipart()
    message['From'] = 'axu307@gmail.com'
    message['To'] =  staff.email
    message['Subject'] = '预约领取统计表'
    if len(row_header) > 4:
        #message.body = "您好，附件中包含了从{}到{}日的预约领取信息。".format(row_header[3], row_header[-1])
        message.attach(MIMEText("您好，附件中包含了从{}到{}日的预约领取信息。".format(row_header[4], row_header[-1]), 'plain'))
    else:
        #message.body = "您好，附件中包含了一些预约领取信息。"
        message.attach(MIMEText("您好，附件中包含了一些预约领取信息。"))
    #message.attach("reservation.xlsx", "application/vnd.ms-excel", bytesIO.getvalue())
    #mail.send(message)

    #part = MIMEBase('application', "octet-stream")
    part = MIMEBase('application', "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    #part.set_payload(bytesIO.getvalue(), charset='utf-8')
    part.set_payload(bytesIO.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition',
                    'attachment; filename="reservation.xlsx"')
    #attachement = MIMEApplication(bytesIO.getvalue(), name='checkin.xlsx')
    #attachement['Content-Disposition'] = 'attachment; filename="checkin.xlsx"'

    #message.attach(attachement)
    message.attach(part)
    if send_email(message, staff.email):
        return jsonify({}), 200
    else:
        return jsonify({}), 422 # Unprocessable Content

@api.route('/statistics/recipe/preview', methods=['POST'])
@admin_required
def recipe_records_preview():
    time_period_option = request.json.get('time_period_option')
    try:
        time_period_option = int(time_period_option)
    except Exception as e:
        abort(make_response(jsonify(errcode=STATUS_REQUIRED_ARG_INVALID, message=MESSAGES[STATUS_REQUIRED_ARG_INVALID] % "time_period_option"), 400)) # bad request

    if time_period_option == 1 or time_period_option== 2:
        these_days = get_week_days(1-time_period_option)
        begin_date = these_days[0]
        end_date = these_days[-1]
    else:
        begin_date = request.json.get('begin_date')
        end_date = request.json.get('end_date')
        try:
            begin_date = date.fromisoformat(begin_date)
            end_date = date.fromisoformat(end_date)
        except Exception as e:
            abort(make_response(jsonify(errcode=STATUS_REQUIRED_ARG_INVALID, message=MESSAGES[STATUS_REQUIRED_ARG_INVALID] % "begin date or end date"), 400)) # bad response

    recipes = Recipe.query.filter(Recipe.begin_time.cast(Date) >= begin_date,
                                            Recipe.begin_time.cast(Date) <= end_date).order_by(Recipe.begin_time.desc()).all()

    result = []
    for r in recipes:
        dishes = []
        for rd in r.dishes:
            ds = rd.dish.to_json()
            ds['image'] = rd.dish.images[0].to_json()
            dishes.append(ds)
        recipe = r.to_json()
        recipe['dishes'] = dishes
        result.append(recipe)

    return jsonify(result), 200

@api.route('/statistics/recipe/email', methods=['POST'])
@admin_required
def recipe_records_email():
    pass

def excel_reservation_record(bytesIO, row_header, records):
    wb = Workbook()
    # grab the active worksheet
    ws = wb.active
    ws['A1'] = "从{}日到{}日的预约领取信息".format(records['begin_date'], records['end_date'])
    ws.append(row_header)
    record = records['records']
    print(record)
    for index, record in enumerate(record, start=1):
        print(record)
        a = [index, record['staff']['last_name'] + record['staff']['first_name'], record['staff']['company']['name']]
        for reservation in record['reservations']:
            a.append(reservation['combo']['name'])
            a.append(reservation['occur_time'])
            if reservation['pickup_time']:
                a.append(reservation['pickup_time'])
        ws.append(a)

    last_column = ord('A')+len(row_header)-1
    last_row = len(records['records']) + 2

    tab = Table(displayName="CheckinTable", ref="A2:{}{}".format(chr(last_column), last_row))
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    #ws.append([])
    #last_line = ord(last_line) + 1
    #for summary in summaries:
    #    ws.append(summary)

    wb.save(bytesIO)

def send_email(message: MIMEMultipart, receiver_email):
    sender_email = 'axu307@gmail.com'
    sender_password = 'msunhklhvrdhlfrm'
    #receiver_email = 'warmlab@outlook.com'
    #message = MIMEMultipart()
    #message['From'] = sender_email
    #message['To'] = receiver_email
    #message['Subject'] = "test"
    #message.attach(MIMEText('test', 'plain'))

    try:
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(sender_email, sender_password)
        server.sendmail(sender_email, receiver_email, message.as_bytes())
        server.quit()

        return True
    except Exception as e:
        print(e)
        return False
